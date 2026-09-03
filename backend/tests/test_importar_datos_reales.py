import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

import openpyxl  # noqa: E402

from importar_datos_reales import (  # noqa: E402
    AmbienteImportado,
    BloqueImportado,
    FichaImportada,
    InstructorImportado,
    detectar_inconsistencias,
    generar_seed_sql,
    parsear_fichas,
    parsear_instructores_ambientes,
    parsear_planeacion,
)


def _bloque(**overrides) -> BloqueImportado:
    base = dict(
        codigo_ficha="FICHA-001",
        instructor_sigla_o_email="DC",
        numero_ambiente="101",
        dia="Lunes",
        jornada="Mañana",
        hora_inicio="08:00",
        hora_fin="10:00",
        resultado_codigo="RA-9",
        fila_origen=1,
    )
    base.update(overrides)
    return BloqueImportado(**base)


def test_detecta_instructor_repetido_en_mismo_dia_jornada():
    bloques = [
        _bloque(fila_origen=1),
        _bloque(fila_origen=2, codigo_ficha="FICHA-002", numero_ambiente="102"),
    ]
    problemas = detectar_inconsistencias(bloques)
    assert any("Instructor 'DC' repetido" in p for p in problemas)


def test_detecta_ambiente_repetido_en_mismo_dia_jornada():
    bloques = [
        _bloque(fila_origen=1, instructor_sigla_o_email="DC"),
        _bloque(fila_origen=2, instructor_sigla_o_email="LM", codigo_ficha="FICHA-002"),
    ]
    problemas = detectar_inconsistencias(bloques)
    assert any("Ambiente '101' repetido" in p for p in problemas)


def test_no_reporta_nada_si_no_hay_choques():
    bloques = [
        _bloque(fila_origen=1, dia="Lunes"),
        _bloque(fila_origen=2, dia="Martes", instructor_sigla_o_email="LM", numero_ambiente="102", codigo_ficha="FICHA-002"),
    ]
    problemas = detectar_inconsistencias(bloques)
    assert problemas == []


def test_reporta_ficha_y_ambiente_fuera_del_catalogo():
    bloques = [_bloque(codigo_ficha="FICHA-999", numero_ambiente="999")]
    problemas = detectar_inconsistencias(
        bloques,
        fichas_validas={"FICHA-001"},
        ambientes_validos={"101"},
    )
    assert any("FICHA-999" in p and "no está" in p for p in problemas)
    assert any("999" in p and "no está" in p for p in problemas)


def test_generar_seed_sql_escribe_archivo_revisable(tmp_path):
    ruta_salida = tmp_path / "seed.sql"
    fichas = [FichaImportada(codigo_ficha="FICHA-001", nombre_programa="Tecnología", sede="Sede Norte")]
    bloques = [_bloque()]

    generar_seed_sql(fichas, [], [AmbienteImportado(numero_ambiente="101")], bloques, ruta_salida)

    contenido = ruta_salida.read_text(encoding="utf-8")
    assert "BEGIN;" in contenido
    assert "COMMIT;" in contenido
    assert "INSERT INTO fichas" in contenido
    assert "INSERT INTO horarios" in contenido
    assert "FICHA-001" in contenido
    # No debe intentar aplicar nada — solo generar el archivo.
    assert "psycopg2" not in contenido and "engine.execute" not in contenido


# --- Parsers reales, contra workbooks sintéticos (no el Excel real de la
# coordinación — ese no se sube al repo). Reproducen el layout confirmado
# a mano contra el archivo real el 2026-09-02.


def _workbook_fichas() -> openpyxl.Workbook:
    import datetime as dt

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FICHAS"
    ws.append(["No. FICHAS", None, "FICHA", "FICHA", "NIVEL", "COORDINACIÓN", "TRI", "INCIO ", "FIN LECTIVA", "FIN PRODUCTIVA"])
    ws.append([1, None, 2996161, "7_TRM_2996161_(DM)_DESARROLLO", "TECNÓLOGO", "MERCADEO", 7, dt.datetime(2024, 7, 8), dt.datetime(2026, 4, 6), dt.datetime(2026, 10, 7)])
    ws.append([2, None, None, None, None, None, None, None, None, None])  # fila sin código, debe ignorarse
    return wb


def test_parsear_fichas(tmp_path):
    ruta = tmp_path / "fichas.xlsx"
    _workbook_fichas().save(ruta)

    fichas = parsear_fichas(ruta)

    assert len(fichas) == 1
    assert fichas[0].codigo_ficha == "2996161"
    assert fichas[0].nombre_programa == "MERCADEO"
    assert fichas[0].fecha_inicio_lectiva == "2024-07-08"
    assert fichas[0].fecha_fin_lectiva == "2026-04-06"
    assert fichas[0].fecha_fin_productiva == "2026-10-07"


def _workbook_instructores_ambientes() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LISTA_INSTRUCTORES_AMBIENTES"
    encabezado = [None] * 19
    encabezado[1], encabezado[2], encabezado[3] = "INICIALES", "NOMBRE COMPLETO", None
    ws.append(encabezado)

    fila_instructor = [None] * 19
    fila_instructor[1], fila_instructor[2], fila_instructor[3] = "CB", "CRISTIAN BUITRAGO", "PLANTA"
    ws.append(fila_instructor)

    fila_ambiente = [None] * 19
    fila_ambiente[9], fila_ambiente[10] = 101, "ESTUDIO 101"
    ws.append(fila_ambiente)

    fila_tema = [None] * 19
    fila_tema[16], fila_tema[17], fila_tema[18] = "INT1", "584305 - ELABORAR EL STORYBOARD", 6
    ws.append(fila_tema)

    # Fila que reutiliza las columnas 1/2 para OTRA cosa (el antipatrón real:
    # más abajo en la hoja esas columnas dejan de ser instructores) — no
    # debe colarse como instructor porque col3 no es PLANTA/CONTRATO.
    fila_reusada = [None] * 19
    fila_reusada[1], fila_reusada[2] = "AR", "AMBIENTE ROTATIVO"
    ws.append(fila_reusada)

    return wb


def test_parsear_instructores_ambientes_separa_los_3_catalogos(tmp_path):
    ruta = tmp_path / "lista.xlsx"
    _workbook_instructores_ambientes().save(ruta)

    instructores, ambientes, temas = parsear_instructores_ambientes(ruta)

    assert len(instructores) == 1
    assert instructores[0] == InstructorImportado(nombre="CRISTIAN BUITRAGO", sigla="CB", tipo_contrato="planta")

    assert len(ambientes) == 1
    assert ambientes[0] == AmbienteImportado(numero_ambiente="101", nombre="ESTUDIO 101")

    assert len(temas) == 1
    assert temas[0]["codigo"] == "584305"
    assert temas[0]["horas"] == 6


def _workbook_planeacion() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLANEACION"
    ws.append(["programa", None, None, "lunes", None, None, None, None, None, "martes", None])
    ws.append(["programa", None, "ficha ", "M", "M", "T", "T", "N", "N", "M", "M"])
    ws.append(["TEMAS_", 1, "TEMAS_7_TRM_2996161_(DM)_DESARROLLO", "TDPM16", "TDPM16", None, None, None, None, None, None])
    ws.append(["INSTRUCTOR_", None, "7_TRM_2996161_(DM)_DESARROLLO", "LA", "LA", None, None, None, None, None, None])
    ws.append(["AMBIENTE_.", None, "AMBIENTE_.7_TRM_2996161_(DM)_DESARROLLO", 505, 505, None, None, None, None, None, None])
    # Segunda ficha SIN fila INSTRUCTOR_ (pasa en 2 de 229 fichas reales) —
    # el bloque debe seguir generándose con instructor vacío, no romperse.
    ws.append(["TEMAS_", 2, "TEMAS_7_TRM_3068352_(MM)_DESARROLLO", None, None, "TDPM13", "TDPM13", None, None, None, None])
    ws.append(["AMBIENTE_.", None, "AMBIENTE_.7_TRM_3068352_(MM)_DESARROLLO", None, None, 604, 604, None, None, None, None])
    return wb


def test_parsear_planeacion_despivotea_y_tolera_instructor_faltante(tmp_path):
    ruta = tmp_path / "planeacion.xlsx"
    _workbook_planeacion().save(ruta)

    bloques = parsear_planeacion(ruta)

    bloques_ficha1 = [b for b in bloques if b.codigo_ficha == "2996161"]
    assert len(bloques_ficha1) == 1
    assert bloques_ficha1[0].dia == "Lunes"
    assert bloques_ficha1[0].jornada == "Mañana"
    assert bloques_ficha1[0].instructor_sigla_o_email == "LA"
    assert bloques_ficha1[0].numero_ambiente == "505"
    assert bloques_ficha1[0].resultado_codigo == "TDPM16"

    bloques_ficha2 = [b for b in bloques if b.codigo_ficha == "3068352"]
    assert len(bloques_ficha2) == 1
    assert bloques_ficha2[0].jornada == "Tarde"
    assert bloques_ficha2[0].instructor_sigla_o_email == ""  # sin fila INSTRUCTOR_, no rompe
    assert bloques_ficha2[0].numero_ambiente == "604"
