"""Import de datos reales del Excel de la coordinación (épica E del backlog
"Nuevo alcance" — SCRUM-71 a 76, ver §7.4 de
_Docs/Documentación general/PLAN_INTEGRACION_LOGICA_Y_BD.md).

Archivo fuente esperado: "PROGRAMACIÓN CGMLTI I TRM 2026.xlsx" (la macro
Excel real que hoy usa la coordinación), con 3 hojas relevantes:
  - FICHAS: catálogo de fichas.
  - LISTA_INSTRUCTORES_AMBIENTES: OJO, apila 3 catálogos sin relación real
    entre sí por coincidencia de fila (instructores, ambientes, un tercero
    sin identificar todavía) — separarlos, no asumir que la fila los
    relaciona (hallazgo de §7.1 del plan).
  - PLANEACION: única hoja que trae los horarios ya asignados, en formato
    ancho (columnas por bloque/día) — hay que despivotear a filas.

Estado (2026-09-02): SCRUM-74 (reportar inconsistencias) y SCRUM-75
(generar seed SQL revisable) están implementados abajo contra la
representación intermedia (Ficha/Instructor/Ambiente/Bloque). SCRUM-71/72/73
(el parseo real de cada hoja) son stubs — necesitan el archivo real para
saber los nombres exactos de columnas; están marcados con TODO.

Uso previsto (una vez completados los parsers):
    cd backend
    .venv/bin/python scripts/importar_datos_reales.py \
        --excel "ruta/PROGRAMACIÓN CGMLTI I TRM 2026.xlsx" \
        --salida ../database/seeds/import_nuevo_alcance.sql

No aplica nada directo a la base de datos — solo genera el .sql para
revisión manual, igual que toda migración en este repo (ver
database/migrations/README.md).
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

# ASUNCIÓN pendiente de confirmar con la coordinación: la hoja PLANEACION
# solo dice la jornada (M/T/N) de cada bloque, no la hora exacta — no hay
# rango de horas estándar documentado en REGLAS_DE_NEGOCIO_CONOCIDAS.md ni
# en los seeds. Este mapeo es un supuesto razonable (jornadas SENA típicas),
# NO un dato confirmado — revisar antes de correr el import real (SCRUM-76).
HORAS_POR_JORNADA = {
    "Mañana": ("06:00", "12:00"),
    "Tarde": ("12:00", "18:00"),
    "Noche": ("18:00", "22:00"),
}


# --- Representación intermedia ---------------------------------------------
# Todo lo de acá para abajo (detectar_inconsistencias, generar_seed_sql) es
# independiente de cómo esté armado el Excel: trabaja contra estos objetos,
# no contra pandas. Así que ya se puede probar/usar en cuanto los parsers de
# arriba llenen estas listas.


@dataclass
class FichaImportada:
    codigo_ficha: str
    nombre_programa: str
    sede: str | None = None
    fecha_inicio_lectiva: str | None = None  # ISO "YYYY-MM-DD"
    fecha_fin_lectiva: str | None = None
    fecha_inicio_productiva: str | None = None
    fecha_fin_productiva: str | None = None


@dataclass
class InstructorImportado:
    nombre: str
    email: str | None = None
    sigla: str | None = None  # ej. "DC" — ver usuarios.sigla, SCRUM-80
    tipo_contrato: str | None = None  # "planta" | "contrato"


@dataclass
class AmbienteImportado:
    numero_ambiente: str
    nombre: str | None = None


@dataclass
class BloqueImportado:
    """Una fila ya despivoteada de PLANEACION: un bloque de clase en un
    día+jornada concretos. `dia`/`jornada` en el mismo vocabulario que
    diasDeLaSemana.nombreDia / jornadas.nombreJornada (ver
    backend/app/models/dia_semana.py, jornada.py) para que el seed SQL
    pueda resolverlos contra esas tablas por nombre."""

    codigo_ficha: str
    instructor_sigla_o_email: str
    numero_ambiente: str
    dia: str
    jornada: str
    hora_inicio: str  # "HH:MM"
    hora_fin: str
    resultado_codigo: str | None = None
    fila_origen: int | None = None  # fila del Excel, para trazar el reporte


# --- SCRUM-71/72/73: parseo del Excel real (stubs, bloqueados por archivo) -


def parsear_fichas(ruta_excel: Path) -> list[FichaImportada]:
    """SCRUM-71. Hoja FICHAS: fila 0 es encabezado, datos desde la fila 1.
    Columnas confirmadas contra el archivo real 2026-09-02: col2=código de
    ficha, col5="COORDINACIÓN" (en la práctica trae el área/programa, ej.
    "MERCADEO"), col7=fecha inicio lectiva, col8=fin lectiva, col9=fin
    productiva, col11=sede (casi siempre vacía en el archivo real).
    OJO: la hoja NO trae fecha de inicio de la etapa productiva por
    separado — queda None; si hace falta, calcularla o pedirla aparte.
    Alternativa a evaluar más adelante: la hoja PE-04 (export crudo de
    SOFIA Plus) puede ser mejor fuente para el catálogo — ver §7.1 del plan."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb["FICHAS"]

    fichas: list[FichaImportada] = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        codigo = fila[2]
        if codigo is None:
            continue
        fichas.append(
            FichaImportada(
                codigo_ficha=str(codigo),
                nombre_programa=(fila[5] or "").strip() or "Sin programa",
                sede=(fila[11] or "").strip() or None if len(fila) > 11 else None,
                fecha_inicio_lectiva=fila[7].date().isoformat() if fila[7] else None,
                fecha_fin_lectiva=fila[8].date().isoformat() if fila[8] else None,
                fecha_inicio_productiva=None,
                fecha_fin_productiva=fila[9].date().isoformat() if fila[9] else None,
            )
        )
    return fichas


def parsear_instructores_ambientes(
    ruta_excel: Path,
) -> tuple[list[InstructorImportado], list[AmbienteImportado], list[dict]]:
    """SCRUM-72. LISTA_INSTRUCTORES_AMBIENTES apila 3 catálogos SIN relación
    real por fila (antipatrón confirmado, §7.1) — se separan por un
    criterio propio de cada uno, no por posición de fila:
      - Instructores: col3 (tipo de contrato) vale exactamente "PLANTA" o
        "CONTRATO" — cualquier otra fila NO es un instructor real, aunque
        tenga algo en col1/col2 (esas columnas se reutilizan más abajo en
        la hoja para otra cosa, confirmado contra el archivo real).
      - Ambientes: col9 (número de ambiente) no es None.
      - Temas/resultados: col17 (actividad de formación) no es None —
        formato "<código> - <descripción>", se separa por el primer " - ".
    Las 3 condiciones son independientes: una fila puede cumplir una, "
    varias o ninguna."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb["LISTA_INSTRUCTORES_AMBIENTES"]

    instructores: list[InstructorImportado] = []
    ambientes: list[AmbienteImportado] = []
    temas: list[dict] = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[3] in ("PLANTA", "CONTRATO"):
            instructores.append(
                InstructorImportado(
                    nombre=(fila[2] or "").strip(),
                    sigla=(fila[1] or "").strip() or None,
                    tipo_contrato=fila[3].lower(),
                )
            )

        if len(fila) > 9 and fila[9] is not None:
            ambientes.append(
                AmbienteImportado(
                    numero_ambiente=str(fila[9]),
                    nombre=(fila[10] or "").strip() if len(fila) > 10 and fila[10] else None,
                )
            )

        if len(fila) > 17 and fila[17]:
            codigo, _, descripcion = str(fila[17]).partition(" - ")
            temas.append(
                {
                    "acronimo": fila[16],
                    "codigo": codigo.strip(),
                    "descripcion": descripcion.strip() or str(fila[17]),
                    "horas": fila[18] if len(fila) > 18 else None,
                }
            )

    return instructores, ambientes, temas


_RE_CODIGO_FICHA = re.compile(r"TRM_(\d{5,8})")


def parsear_planeacion(ruta_excel: Path) -> list[BloqueImportado]:
    """SCRUM-73. Hoja PLANEACION, formato ancho confirmado contra el
    archivo real: fila 0 = nombre del día por grupos de 6 columnas
    (lunes..sabado, empezando en la columna índice 3), fila 1 = código de
    jornada por columna (M/M/T/T/N/N — 2 sub-bloques por jornada, se
    tratan como un solo bloque día×jornada, igual que hace la macro real
    y consistente con HorarioService._detectar_cruces, ver §7.1 del plan).

    Cada ficha ocupa un grupo de filas con etiqueta en la columna 0:
    "TEMAS_" (tema por bloque), luego opcionalmente "INSTRUCTOR_" (a veces
    falta en el archivo real — 227 de 229 fichas la traen) y
    "AMBIENTE_." (siempre presente, cierra el grupo de la ficha). El
    código de ficha se extrae con regex del texto de la fila TEMAS_
    ("TEMAS_7_TRM_2996161_..." → "2996161"), que coincide con el código
    de la hoja FICHAS."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb["PLANEACION"]
    filas = list(ws.iter_rows(min_row=1, values_only=True))

    fila_dias = filas[0]
    fila_jornadas = filas[1]

    # (día, jornada) -> lista de columnas (2 sub-bloques por jornada en el
    # archivo real) — se colapsan a UN bloque por jornada, tomando el
    # primer sub-bloque no vacío, no uno por cada columna.
    bloque_a_columnas: dict[tuple[str, str], list[int]] = defaultdict(list)
    dia_actual = None
    for col_idx in range(3, len(fila_dias)):
        if fila_dias[col_idx]:
            dia_actual = str(fila_dias[col_idx]).strip().capitalize()
        codigo_jornada = fila_jornadas[col_idx] if col_idx < len(fila_jornadas) else None
        jornada = {"M": "Mañana", "T": "Tarde", "N": "Noche"}.get(codigo_jornada)
        if dia_actual and jornada:
            bloque_a_columnas[(dia_actual, jornada)].append(col_idx)

    bloques: list[BloqueImportado] = []
    grupo: dict | None = None

    def cerrar_grupo():
        if grupo is None or grupo["codigo_ficha"] is None:
            return
        fila_temas = grupo["temas"]
        fila_instructor = grupo["instructor"]
        fila_ambiente = grupo["ambiente"]

        def _primer_valor(fila, columnas):
            if not fila:
                return None
            for c in columnas:
                if c < len(fila) and fila[c] is not None:
                    return fila[c]
            return None

        for (dia, jornada), columnas in bloque_a_columnas.items():
            tema = _primer_valor(fila_temas, columnas)
            instructor = _primer_valor(fila_instructor, columnas)
            ambiente = _primer_valor(fila_ambiente, columnas)
            if tema is None and instructor is None and ambiente is None:
                continue
            hora_inicio, hora_fin = HORAS_POR_JORNADA[jornada]
            bloques.append(
                BloqueImportado(
                    codigo_ficha=grupo["codigo_ficha"],
                    instructor_sigla_o_email=str(instructor) if instructor is not None else "",
                    numero_ambiente=str(ambiente) if ambiente is not None else "",
                    dia=dia,
                    jornada=jornada,
                    hora_inicio=hora_inicio,
                    hora_fin=hora_fin,
                    resultado_codigo=str(tema) if tema is not None else None,
                    fila_origen=grupo["fila_origen"],
                )
            )

    for num_fila, fila in enumerate(filas[2:], start=3):
        etiqueta = fila[0]
        if etiqueta == "TEMAS_":
            cerrar_grupo()
            match = _RE_CODIGO_FICHA.search(str(fila[2] or ""))
            grupo = {
                "codigo_ficha": match.group(1) if match else None,
                "temas": fila,
                "instructor": None,
                "ambiente": None,
                "fila_origen": num_fila,
            }
        elif etiqueta == "INSTRUCTOR_" and grupo is not None:
            grupo["instructor"] = fila
        elif etiqueta == "AMBIENTE_." and grupo is not None:
            grupo["ambiente"] = fila

    cerrar_grupo()
    return bloques


# --- SCRUM-74: reportar cruces/inconsistencias, sin insertar en silencio ---


def detectar_inconsistencias(
    bloques: list[BloqueImportado],
    fichas_validas: set[str] | None = None,
    ambientes_validos: set[str] | None = None,
) -> list[str]:
    """Cruces a nivel de bloque fijo (día×jornada) — mismo criterio que usa
    la macro Excel real (§7.1 del plan): un instructor o un ambiente no
    puede repetirse en el mismo día+jornada en dos bloques distintos.
    También reporta referencias a fichas/ambientes que no están en el
    catálogo ya parseado, si se pasan `fichas_validas`/`ambientes_validos`.
    No modifica nada — solo arma el reporte para que la coordinación lo
    revise antes de que se genere o aplique cualquier seed."""
    problemas: list[str] = []

    por_instructor_bloque: dict[tuple[str, str, str], list[BloqueImportado]] = defaultdict(list)
    por_ambiente_bloque: dict[tuple[str, str, str], list[BloqueImportado]] = defaultdict(list)

    for bloque in bloques:
        clave_instructor = (bloque.instructor_sigla_o_email, bloque.dia, bloque.jornada)
        por_instructor_bloque[clave_instructor].append(bloque)

        clave_ambiente = (bloque.numero_ambiente, bloque.dia, bloque.jornada)
        por_ambiente_bloque[clave_ambiente].append(bloque)

        if fichas_validas is not None and bloque.codigo_ficha not in fichas_validas:
            problemas.append(
                f"Fila {bloque.fila_origen}: la ficha '{bloque.codigo_ficha}' no está "
                "en el catálogo de FICHAS parseado."
            )
        if ambientes_validos is not None and bloque.numero_ambiente not in ambientes_validos:
            problemas.append(
                f"Fila {bloque.fila_origen}: el ambiente '{bloque.numero_ambiente}' no "
                "está en el catálogo de ambientes parseado."
            )

    for (instructor, dia, jornada), choques in por_instructor_bloque.items():
        if len(choques) > 1:
            filas = ", ".join(str(b.fila_origen) for b in choques)
            problemas.append(
                f"Instructor '{instructor}' repetido en {dia} {jornada} — filas: {filas}."
            )

    for (ambiente, dia, jornada), choques in por_ambiente_bloque.items():
        if len(choques) > 1:
            filas = ", ".join(str(b.fila_origen) for b in choques)
            problemas.append(
                f"Ambiente '{ambiente}' repetido en {dia} {jornada} — filas: {filas}."
            )

    return problemas


# --- SCRUM-75: generar SQL/CSV de seed revisable, sin aplicar directo -----


def _escapar(valor: str | None) -> str:
    if valor is None:
        return "NULL"
    return "'" + valor.replace("'", "''") + "'"


def generar_seed_sql(
    fichas: list[FichaImportada],
    instructores: list[InstructorImportado],
    ambientes: list[AmbienteImportado],
    bloques: list[BloqueImportado],
    ruta_salida: Path,
) -> None:
    """Escribe un .sql revisable en `ruta_salida` — NO se aplica acá ni se
    llama a ningún engine/conexión. El usuario lo revisa y lo corre
    manualmente (mismo patrón que database/migrations/, ver
    feedback_supabase_schema_changes en la memoria del proyecto).

    Los instructores/ambientes se resuelven contra las tablas existentes
    por nombre/sigla (subselect), asumiendo que ya fueron cargados antes
    (por ejemplo por SCRUM-77/80, ya mergeados) — este script no inventa
    UUIDs de Supabase Auth para instructores nuevos, eso requiere pasar por
    el flujo normal de creación de usuario."""

    lineas: list[str] = [
        "-- Generado por backend/scripts/importar_datos_reales.py — REVISAR ANTES DE CORRER.",
        "-- No se aplica automáticamente. Ejecutar manualmente contra un ambiente de prueba primero (SCRUM-76).",
        "BEGIN;",
        "",
    ]

    if fichas:
        lineas.append("-- Fichas (upsert por codigoFicha)")
        for f in fichas:
            lineas.append(
                "INSERT INTO fichas (\"codigoFicha\", \"idSede\", \"fechaInicioLectiva\", "
                '"fechaFinLectiva", "fechaInicioProductiva", "fechaFinProductiva") '
                f"VALUES ({_escapar(f.codigo_ficha)}, "
                f"(SELECT \"idSede\" FROM sedes WHERE \"nombreSede\" = {_escapar(f.sede)}), "
                f"{_escapar(f.fecha_inicio_lectiva)}, {_escapar(f.fecha_fin_lectiva)}, "
                f"{_escapar(f.fecha_inicio_productiva)}, {_escapar(f.fecha_fin_productiva)}) "
                'ON CONFLICT ("codigoFicha") DO UPDATE SET '
                '"idSede" = EXCLUDED."idSede", "fechaInicioLectiva" = EXCLUDED."fechaInicioLectiva", '
                '"fechaFinLectiva" = EXCLUDED."fechaFinLectiva", '
                '"fechaInicioProductiva" = EXCLUDED."fechaInicioProductiva", '
                '"fechaFinProductiva" = EXCLUDED."fechaFinProductiva";'
            )
        lineas.append("")

    if bloques:
        lineas.append("-- Horarios importados de PLANEACION")
        lineas.append(
            "-- (usa subselects por sigla/número de ambiente/código de ficha; si alguno no "
            "existe, el INSERT falla ruidosamente en vez de guardar NULL en silencio)"
        )
        for b in bloques:
            lineas.append(
                "INSERT INTO horarios (\"horaInicio\", \"horaFin\", \"idJornada\", \"idAmbiente\", "
                '"idInstructor", "idFicha", "idResultado") VALUES ('
                f"{_escapar(b.hora_inicio)}, {_escapar(b.hora_fin)}, "
                f"(SELECT \"idJornada\" FROM jornadas WHERE \"nombreJornada\" = {_escapar(b.jornada)}), "
                f"(SELECT \"idAmbiente\" FROM ambientes WHERE \"numeroAmbiente\" = {_escapar(b.numero_ambiente)}), "
                f"(SELECT \"idUsuario\" FROM usuarios WHERE sigla = {_escapar(b.instructor_sigla_o_email)} "
                f"OR email = {_escapar(b.instructor_sigla_o_email)}), "
                f"(SELECT \"idFicha\" FROM fichas WHERE \"codigoFicha\" = {_escapar(b.codigo_ficha)}), "
                f"(SELECT \"idResultado\" FROM resultados_aprendizaje WHERE codigo = {_escapar(b.resultado_codigo)})"
                f"); -- fila origen Excel: {b.fila_origen}"
            )
        lineas.append("")

    lineas.append("COMMIT;")

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    ruta_salida.write_text("\n".join(lineas), encoding="utf-8")


# --- CLI ---------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, help="Ruta al .xlsx real de la coordinación")
    parser.add_argument(
        "--salida",
        default="../database/seeds/import_nuevo_alcance.sql",
        help="Dónde escribir el .sql revisable (default: database/seeds/import_nuevo_alcance.sql)",
    )
    args = parser.parse_args()

    ruta_excel = Path(args.excel)
    fichas = parsear_fichas(ruta_excel)
    instructores, ambientes, _otro_catalogo = parsear_instructores_ambientes(ruta_excel)
    bloques = parsear_planeacion(ruta_excel)

    problemas = detectar_inconsistencias(
        bloques,
        fichas_validas={f.codigo_ficha for f in fichas},
        ambientes_validos={a.numero_ambiente for a in ambientes},
    )
    if problemas:
        print(f"Se encontraron {len(problemas)} inconsistencias — revisar antes de generar el seed:\n")
        for p in problemas:
            print(f"  - {p}")
        print("\nNo se generó el .sql. Corrige los datos o filtra los bloques problemáticos y vuelve a correr.")
        sys.exit(1)

    ruta_salida = Path(args.salida)
    generar_seed_sql(fichas, instructores, ambientes, bloques, ruta_salida)
    print(f"Seed generado en {ruta_salida} — revísalo y córrelo manualmente contra el ambiente de prueba.")


if __name__ == "__main__":
    main()
